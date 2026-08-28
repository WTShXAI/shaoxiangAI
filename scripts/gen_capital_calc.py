# -*- coding: utf-8 -*-
"""Generate 资金方案计算器.xlsx with live Excel formulas.
Fill yellow input cells; computed cells update automatically in Excel."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

OUT = r"D:/Architecture/资金方案计算器.xlsx"

wb = openpyxl.Workbook()

# ---------- styles ----------
title_font = Font(name="微软雅黑", size=14, bold=True, color="1F3864")
hdr_font   = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
lbl_font   = Font(name="微软雅黑", size=10)
in_font    = Font(name="微软雅黑", size=10, bold=True, color="7F6000")
out_font   = Font(name="微软雅黑", size=10, bold=True, color="1F3864")

fill_title = PatternFill("solid", fgColor="DDEBF7")
fill_in    = PatternFill("solid", fgColor="FFF2CC")   # yellow input
fill_out   = PatternFill("solid", fgColor="E2EFDA")   # green computed
fill_hdr   = PatternFill("solid", fgColor="4472C4")
fill_sec   = PatternFill("solid", fgColor="BDD7EE")

thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

MONEY = '#,##0'
PCT   = '0.0%'

def style_label(c):
    c.font = lbl_font; c.border = border
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

def style_input(c, fmt=None):
    c.font = in_font; c.fill = fill_in; c.border = border
    c.alignment = Alignment(horizontal="right", vertical="center")
    if fmt: c.number_format = fmt

def style_output(c, fmt=None):
    c.font = out_font; c.fill = fill_out; c.border = border
    c.alignment = Alignment(horizontal="right", vertical="center")
    if fmt: c.number_format = fmt

# ============================================================
# Sheet 1: 资金方案计算器
# ============================================================
ws = wb.active
ws.title = "资金方案计算器"
ws.sheet_view.showGridLines = False

ws.merge_cells("A1:C1")
ws["A1"] = "哨响AI · 资金方案计算器（填黄色格，结果自动算）"
ws["A1"].font = title_font; ws["A1"].fill = fill_title
ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
ws.row_dimensions[1].height = 24

ws.merge_cells("A2:C2")
ws["A2"] = "原则：作战资金=总盘×比例；每注=作战×f 再封顶；回撤/日亏/周亏触发止损；季度再平衡拉回比例。"
ws["A2"].font = Font(name="微软雅黑", size=9, italic=True, color="595959")
ws["A2"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
ws.row_dimensions[2].height = 28

# --- Inputs (rows 4-12) ---
inputs = [
    (4,  "总盘（元）",            1000000, MONEY),
    (5,  "作战资金比例",          0.40,    PCT),
    (6,  "注码比例 f（每注占比）",0.015,   PCT),
    (7,  "单注绝对上限（元）",    10000,   MONEY),
    (8,  "最大回撤止损（%）",     0.25,    PCT),
    (9,  "单日亏损上限（%）",     0.05,    PCT),
    (10, "单周亏损上限（%）",     0.10,    PCT),
    (11, "并发注数上限（笔）",    5,       '0'),
]
for r, name, val, fmt in inputs:
    ws.cell(r, 1, name); style_label(ws.cell(r, 1))
    c = ws.cell(r, 2, val); style_input(c, fmt)
    ws.cell(r, 3, "输入"); ws.cell(r, 3).font = Font(name="微软雅黑", size=9, color="BF8F00")
    ws.cell(r, 3).alignment = Alignment(horizontal="center")

# reserve ratio auto = 1 - active
ws.cell(12, 1, "储备比例（=1-作战）"); style_label(ws.cell(12, 1))
ws.cell(12, 2, "=1-B5"); style_input(ws.cell(12, 2), PCT)
ws.cell(12, 3, "自动"); ws.cell(12, 3).font = Font(name="微软雅黑", size=9, color="548235")
ws.cell(12, 3).alignment = Alignment(horizontal="center")

# --- Section: 计算结果 (row 14 header) ---
ws.merge_cells("A14:C14")
ws["A14"] = "计算结果（自动）"
ws["A14"].font = hdr_font; ws["A14"].fill = fill_hdr
ws["A14"].alignment = Alignment(horizontal="left", vertical="center")

calc = [
    (15, "作战资金（元）",        "=B4*B5",            MONEY),
    (16, "储备资金（元）",        "=B4*B12",           MONEY),
    (17, "基础单位（元）",        "=B15*B6",           MONEY),
    (18, "实际单注（封顶后，元）","=MIN(B17,B7)",      MONEY),
    (19, "回撤止损线（作战，元）","=B15*(1-B8)",       MONEY),
    (20, "单日亏损上限（元）",    "=B15*B9",           MONEY),
    (21, "单周亏损上限（元）",    "=B15*B10",          MONEY),
    (22, "满仓并发最大敞口（元）","=B18*B11",          MONEY),
]
for r, name, formula, fmt in calc:
    ws.cell(r, 1, name); style_label(ws.cell(r, 1))
    c = ws.cell(r, 2, formula); style_output(c, fmt)
    ws.cell(r, 3, "结果"); ws.cell(r, 3).font = Font(name="微软雅黑", size=9, color="548235")
    ws.cell(r, 3).alignment = Alignment(horizontal="center")

# --- Section: 分阶段放量单位 (row 24 header) ---
ws.merge_cells("A24:C24")
ws["A24"] = "分阶段放量单位（作战资金 × 阶段 f）"
ws["A24"].font = hdr_font; ws["A24"].fill = fill_hdr
ws["A24"].alignment = Alignment(horizontal="left", vertical="center")

phase = [
    (25, "阶段0 纸面/微注 (f=0.5%)", "=B15*0.005", MONEY),
    (26, "阶段1 放量     (f=1.5%)", "=B15*0.015", MONEY),
    (27, "阶段2 满仓     (f=2.0%)", "=B15*0.020", MONEY),
]
for r, name, formula, fmt in phase:
    ws.cell(r, 1, name); style_label(ws.cell(r, 1))
    c = ws.cell(r, 2, formula); style_output(c, fmt)
    ws.cell(r, 3, "结果"); ws.cell(r, 3).font = Font(name="微软雅黑", size=9, color="548235")
    ws.cell(r, 3).alignment = Alignment(horizontal="center")

# --- Section: 再平衡模拟 (row 29 header) ---
ws.merge_cells("A29:C29")
ws["A29"] = "再平衡模拟（连赢后把超额扫回储备）"
ws["A29"].font = hdr_font; ws["A29"].fill = fill_hdr
ws["A29"].alignment = Alignment(horizontal="left", vertical="center")

rb = [
    (30, "再平衡时总盘（元）",      1120000, MONEY, "输入"),
    (31, "目标作战资金（元）",      "=B30*B5", MONEY, "结果"),
    (32, "当前作战资金（元）",      520000,  MONEY, "输入"),
    (33, "超额/缺口（正=扫储备，负=补作战，元）", "=B31-B32", MONEY, "结果"),
]
for r, name, val, fmt, tag in rb:
    ws.cell(r, 1, name); style_label(ws.cell(r, 1))
    if tag == "输入":
        c = ws.cell(r, 2, val); style_input(c, fmt)
    else:
        c = ws.cell(r, 2, val); style_output(c, fmt)
    ws.cell(r, 3, tag); ws.cell(r, 3).font = Font(name="微软雅黑", size=9,
        color=("BF8F00" if tag=="输入" else "548235"))
    ws.cell(r, 3).alignment = Alignment(horizontal="center")

ws.column_dimensions["A"].width = 34
ws.column_dimensions["B"].width = 18
ws.column_dimensions["C"].width = 8

# ============================================================
# Sheet 2: 使用说明
# ============================================================
ws2 = wb.create_sheet("使用说明")
ws2.sheet_view.showGridLines = False
ws2.merge_cells("A1:B1")
ws2["A1"] = "资金方案方法说明"
ws2["A1"].font = title_font; ws2["A1"].fill = fill_title
ws2.row_dimensions[1].height = 24

notes = [
    ("1. 资金分层", "总盘=可承受归零的钱；作战资金=唯一下注来源（建议40%）；储备=回撤补给/利润再投（建议60%）。"),
    ("2. 注码方案", "每注=作战×f（推荐1%-2%），等权 flat，禁 Kelly 变量注码；单注硬封顶=min(单位,绝对上限)。"),
    ("3. 复利机制", "作战资金内部随盈利自动变大→下注单位自动变大（复利引擎）；季度再平衡拉回40/60比例，防过度杠杆。"),
    ("4. 风险闸门", "作战回撤≥25% / 单日亏≥5% / 单周亏≥10% → 强制停手复盘，不摊平。"),
    ("5. 并发敞口", "满仓并发最大敞口=实际单注×并发注数上限，避免同联赛 correlated 连环黑。"),
    ("6. 分阶段放量", "阶段0微注验证live ROI→阶段1放量→阶段2满仓（≥200-500注确认edge稳定）。"),
    ("7. 铁律", "绝不在单注/单日维度因连赢加注、因连黑减注（追逐方差）；只在预定再平衡点调整。"),
    ("颜色约定", "黄色=输入格（你填）；绿色=结果格（公式自动算）；储备比例与再平衡目标为自动公式。"),
]
r = 3
for k, v in notes:
    ws2.cell(r, 1, k).font = Font(name="微软雅黑", size=10, bold=True, color="1F3864")
    ws2.cell(r, 1).alignment = Alignment(vertical="top", wrap_text=True)
    ws2.cell(r, 1).border = border
    ws2.cell(r, 2, v).font = Font(name="微软雅黑", size=10)
    ws2.cell(r, 2).alignment = Alignment(vertical="top", wrap_text=True)
    ws2.cell(r, 2).border = border
    ws2.row_dimensions[r].height = 42
    r += 1
ws2.column_dimensions["A"].width = 16
ws2.column_dimensions["B"].width = 78

wb.save(OUT)
print("saved:", OUT)
