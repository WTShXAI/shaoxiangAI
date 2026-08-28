#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
将 events.db 采集到的赔率 + 赛果, 渲染成用户模板(footballct.CSV / .xls)的 Excel 布局.

模板逐格结构 (0-indexed 列号, 实测自 footballct.CSV):
  盘口块:
    A[0]  联赛名(仅头行) / 空(数据行)
    B[1]  队名(主队/客队) / "初盘赔率"(头行)
    C[2]  (保留; 原半场比分已移至 K 列)
    D[3]  全场独赢
    E[4]  全场让球
    F[5]  全场大小
    G[6]  (保留; 原终场比分已移至 L 列)
    H[7]  下半场独赢
    I[8]  下半场让球
    J[9]  下半场大小
    K[10] 半场比分  (格式: 该行队伍进球整数, 如 2; 无则 0)
    L[11] 终场比分  (格式同半场比分; 无则 0)

  波胆块(每场比赛固定渲染, 无波胆则填 '--'):
    标题行: B=初盘波胆比分, E=上半场波胆, H=下半场波胆
    子表头行: B=主队名, C=全场平局, D=客队名, E=主队名, F=上半场平局, G=客队名, H=主队名, I=下半场平局, J=客队名
    数据行:
      B/C/D=初盘全场波胆(主/平/客): GQ 当前无独立全场 CS 初盘, 留 '--'.
      E/F/G=上半场波胆: GQ 无 CS_1H 市场, 留 '--'.
      H/I/J=下半场波胆: 使用 GQ 的 CS 市场数据(用户确认其实际为下半场/剩余时间波胆).

赔率格式:
  主胜2.06 / 客胜2.94 / 和局3.20
  让-0/0.5 2.03 / 让+0/0.5 1.83
  大2.5 1.99 / 小2.5 1.85
  无赔率时保留标签: 主胜-- / 让-- / 大--

波胆格式:
  1:0      8.50   (比分 + 空格 + 赔率)
  0:1      --     (无赔率)
  其他     35.00
  该格不存在(如平局列后4行) -> 空

命中红字 (均为"比赛正确结果"的标注, 有终场比分即标, 不依赖赔率是否存在):
  全场波胆格 == 终场比分 时 -> 红色字体
  上半场波胆格 == 半场比分 时 -> 红色字体 (GQ 无半场波胆, 通常不会触发)
  下半场波胆格 == 下半场比分 (若未来有) 时 -> 红色字体
  全场1X2 (主胜/和局/客胜) -> 按终场胜负标红对应行(全场独赢列 D); 和局标和局行
  终场比分 (主客各自行 G 列) -> 直接标红, 高亮"结果比分"
  全场让球列 -> 终场比分判定赢盘方向(主队行=上盘/客队行=下盘) -> 红字; 走水不标
  全场大小球列 -> 终场总进球判定大/小(主队行=大/客队行=小) -> 红字; 走水不标

数据来源:
  - 全场盘口/波胆: odds_snapshots(非 _2H/_1H) 取最早非零快照=初盘.
  - 下半场盘口: odds_snapshots(_2H) — 仅比赛进行中(下半场)时 GQ 返回.
  - 半场比分: matches/match_outcomes 的 ht_score_home/away(乐鱼 live 采集或历史回填);
        缺失且为 live 比赛时退化为实时比分(近似半场); 其余 0.0. 有实测/实时半场即标红.
  - 终场比分: match_outcomes(经 matches.mid join) 的 score_home/score_away.
  - GQ 无半场波胆市场 -> 上半场/下半场波胆格填 '--'.

用法:
  python scripts/export_to_template.py                      # 全量
  python scripts/export_to_template.py --league "中超"       # 单联赛(模糊)
  python scripts/export_to_template.py --limit 50            # 限制场数
"""
import sqlite3, math, argparse, os
from openpyxl.styles import Font

DB = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "data", "events.db")
OUT = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "data", "export_template.xlsx")

# 模板波胆比分顺序 (主/平/客 三元组, None=该格空)
CS_ROWS = [
    ("1:0", "0:0", "0:1"),
    ("2:0", "1:1", "0:2"),
    ("2:1", "2:2", "0:3"),
    ("3:0", "3:3", "0:4"),
    ("3:1", "4:4", "1:2"),
    ("3:2", "其他", "1:3"),
    ("4:0", None,  "1:4"),
    ("4:1", None,  "2:3"),
    ("4:2", None,  "2:4"),
    ("4:3", None,  "3:4"),
]


# ---------------- 格式化 ----------------
def _num(v):
    return str(int(v)) if float(v) == int(v) else f"{v}"


def fmt_ah_line(lv):
    """亚洲盘口线 -> 标准记法(正负号都正确): -0.25->'0/-0.5', 3.75->'3.5/4'."""
    if lv is None:
        return ""
    if lv == 0:
        return "0"
    sign = -1 if lv < 0 else 1
    a = abs(lv)
    n = int(math.floor(a))
    frac = a - n
    if frac == 0:
        return _num(sign * n)
    if frac == 0.5:
        return _num(sign * (n + 0.5))
    if frac == 0.25:
        return f"{_num(sign * n)}/{_num(sign * (n + 0.5))}"
    if frac == 0.75:
        return f"{_num(sign * (n + 0.5))}/{_num(sign * (n + 1))}"
    return _num(sign * a)


def fmt_1x2(sel, odds=None):
    """无赔率时保留标签: 主胜-- / 和局-- / 客胜--"""
    name = {"home": "主胜", "draw": "和局", "away": "客胜"}.get(sel, sel)
    if odds is None or odds <= 0:
        return f"{name}--"
    return f"{name}{odds:.2f}"


def fmt_ah(sel, lv, odds=None):
    """让-- / 让-0/0.5     2.03 (6空格对齐, 同用户模板)"""
    if odds is None or odds <= 0 or lv is None:
        return "让--"
    return f"让{fmt_ah_line(lv)}      {odds:.2f}"


def fmt_ou(sel, lv, odds=None):
    """大-- / 小-- / 大2.5     1.99 (6空格对齐, 同用户模板)"""
    if odds is None or odds <= 0 or lv is None:
        return f"{'大' if sel == 'over' else '小'}--"
    return f"{'大' if sel == 'over' else '小'}{fmt_ah_line(lv)}      {odds:.2f}"


def fmt_score_line(my, opp):
    """比分渲染: 主队行=主.客, 客队行=客.主; 无则 0.0."""
    if my is None or opp is None:
        return "0.0"
    return f"{int(my)}:{int(opp)}"


def fmt_goal(v):
    """该行队伍进球数渲染: 2 -> '2'; 无则 '0'. 比分列用整数,不带小数点."""
    if v is None:
        return "0"
    return str(int(v))


# ---------------- 波胆 selection 解析 ----------------
def parse_cs_selection(sel):
    """GQ 波胆 selection 两种形态: '1-0' / 'home/1-0' / 'home/home_other' 等.
    归一为 (side, score), side∈home/draw/away, score 形如 '1-0' 或 '其他'."""
    if not sel:
        return None
    s = str(sel).strip()
    side = None
    body = s
    if "/" in s:
        p, body = s.split("/", 1)
        if p in ("home", "draw", "away"):
            side = p
    body = body.strip()
    if body in ("其他", "other") or body.endswith("_other"):
        return (side or "draw", "其他")
    if "-" in body:
        try:
            a, b = body.split("-", 1)
            a, b = int(a), int(b)
            if side is None:
                side = "home" if a > b else ("away" if a < b else "draw")
            return (side, f"{a}-{b}")
        except ValueError:
            return None
    return None


def _line_from_mkt(mkt):
    try:
        return float(mkt.rsplit("_", 1)[1])
    except (IndexError, ValueError):
        return None


# ---------------- 数据加载 ----------------
def load_opening(cur):
    """全场(非半场)市场 + 波胆. 取最早非零快照=初盘."""
    sql = """
    SELECT match_key, market, selection, odds FROM (
      SELECT match_key, market, selection, odds,
             ROW_NUMBER() OVER (PARTITION BY match_key, market, selection
                                ORDER BY (CASE WHEN odds>0 THEN 0 ELSE 1 END), captured_at ASC) rn
      FROM odds_snapshots
      WHERE market NOT LIKE '%_2H%' AND market NOT LIKE '%_1H%'
    ) WHERE rn=1
    """
    ft, cs = {}, {}
    ah_cnt, ou_cnt = {}, {}
    for mk, mkt, sel, odds in cur.execute(sql):
        if mkt == "CS":
            p = parse_cs_selection(sel)
            if p:
                cs[(mk, p[0], p[1])] = odds
            continue
        ft[(mk, mkt, sel)] = odds
        if mkt.startswith("AH_"):
            lv = _line_from_mkt(mkt)
            if lv is not None:
                ah_cnt.setdefault(mk, {}).setdefault(lv, 0)
                ah_cnt[mk][lv] += 1
        elif mkt.startswith("OU_"):
            lv = _line_from_mkt(mkt)
            if lv is not None:
                ou_cnt.setdefault(mk, {}).setdefault(lv, 0)
                ou_cnt[mk][lv] += 1
    ft_ah_line = {mk: max(d, key=d.get) for mk, d in ah_cnt.items() if d}
    ft_ou_line = {mk: max(d, key=d.get) for mk, d in ou_cnt.items() if d}
    return ft, cs, ft_ah_line, ft_ou_line


def load_half(cur):
    """下半场/上半场市场 + 半场波胆(CS_1H/CS_2H)."""
    sql = """
    SELECT match_key, market, selection, odds, line FROM (
      SELECT match_key, market, selection, odds, line,
             ROW_NUMBER() OVER (PARTITION BY match_key, market, selection
                                ORDER BY (CASE WHEN odds>0 THEN 0 ELSE 1 END), captured_at ASC) rn
      FROM odds_snapshots
      WHERE market LIKE '%_2H%' OR market LIKE '%_1H%' OR market IN ('CS_1H','CS_2H')
    ) WHERE rn=1
    """
    h1x2, hah, hou, hcs = {}, {}, {}, {}
    hah_cnt, hou_cnt = {}, {}
    for mk, mkt, sel, odds, line in cur.execute(sql):
        if mkt.startswith("1X2_"):
            suf = "_2H" if "2H" in mkt else "_1H"
            h1x2.setdefault((mk, suf), {})[sel] = odds
        elif mkt.startswith("AH_"):
            suf = "_2H" if "2H" in mkt else "_1H"
            lv = line if line is not None else _line_from_mkt(mkt)
            if lv is not None:
                hah.setdefault((mk, suf, lv), {})[sel] = odds
                hah_cnt.setdefault((mk, suf), {}).setdefault(lv, 0)
                hah_cnt[(mk, suf)][lv] += 1
        elif mkt.startswith("OU_"):
            suf = "_2H" if "2H" in mkt else "_1H"
            lv = line if line is not None else _line_from_mkt(mkt)
            if lv is not None:
                hou.setdefault((mk, suf, lv), {})[sel] = odds
                hou_cnt.setdefault((mk, suf), {}).setdefault(lv, 0)
                hou_cnt[(mk, suf)][lv] += 1
        elif mkt in ("CS_1H", "CS_2H"):
            suf = "_2H" if "2H" in mkt else "_1H"
            p = parse_cs_selection(sel)
            if p:
                hcs[(mk, suf, p[0], p[1])] = odds
    hah_line = {(mk, suf): max(d, key=d.get) for (mk, suf), d in hah_cnt.items() if d}
    hou_line = {(mk, suf): max(d, key=d.get) for (mk, suf), d in hou_cnt.items() if d}
    return h1x2, hah, hou, hcs, hah_line, hou_line


def _cs_cell(cs, mk, side, template_score):
    if not template_score:
        return ""
    score = "其他" if template_score == "其他" else template_score.replace(":", "-")
    odds = cs.get((mk, side, score))
    if not odds or odds <= 0:
        return f"{template_score}      --"
    return f"{template_score}      {odds:.2f}"


def _cs_hit_label(side, template_score):
    """把 (side, template_score) 转成 '1:0' / '其他' 等命中标签, 用于和比分字符串比较."""
    if not template_score:
        return None
    return template_score


# ---------------- 渲染 ----------------
def render_match_block(ws, r, mk, home, away, status, m_h, m_a, ht_h, ht_a,
                       has_outcome, sh, sa,
                       ft, cs, ft_ah_line, ft_ou_line,
                       h1x2, hah, hou, hcs, hah_line, hou_line):
    def C(col0, val):
        ws.cell(r, col0 + 1, val)

    red = Font(color="FF0000")

    # 半场比分 — 只认确认型 (ht_score_home/away: 乐鱼 live 采集 或 历史回填).
    # 曾用「live 比赛当前实时比分」近似半场, 已废弃: 20' 的 0-0 被写进"半场比分"列
    # 属于把未知伪装成已知, 违背「有据可查」. 未采到一律 '--', 覆盖率随采集器积累上升.
    if ht_h is not None and ht_a is not None:
        ht_home = fmt_goal(ht_h)
        ht_away = fmt_goal(ht_a)
        ht_hit = f"{int(ht_h)}:{int(ht_a)}"
        has_ht = True
        ht_confirmed = True      # 实测/回填半场 -> 可作为模型特征标红
    else:
        # 未采到半场比分 -> 留 '--'.
        # 严禁填 0: 0 是合法比分, 填 0 会把"未知"伪装成"实际 0 球", 污染人工研判与训练标注.
        ht_home = ht_away = "--"
        ht_hit = None
        has_ht = False
        ht_confirmed = False
    if has_outcome:
        ft_home = fmt_goal(sh)
        ft_away = fmt_goal(sa)
        ft_hit = f"{int(sh)}:{int(sa)}"    # 用于和波胆格 "x:y" 比较
    else:
        # 未完场/无赛果 -> 留 '--' (同上, 不可填 0)
        ft_home = ft_away = "--"
        ft_hit = None

    # 主线
    ah_lv = ft_ah_line.get(mk)
    ou_lv = ft_ou_line.get(mk)
    ah2_lv = hah_line.get((mk, "_2H"))
    ou2_lv = hou_line.get((mk, "_2H"))

    def g_ft_x2(sel):
        return ft.get((mk, "1X2", sel))

    def g_ft_ah(sel):
        if ah_lv is not None:
            return ft.get((mk, f"AH_{ah_lv:.2f}", sel))
        return None

    def g_ft_ou(sel):
        if ou_lv is not None:
            return ft.get((mk, f"OU_{ou_lv:.2f}", sel))
        return None

    def g_2h_x2(sel):
        return h1x2.get((mk, "_2H"), {}).get(sel)

    def g_2h_ah(sel):
        return hah.get((mk, "_2H", ah2_lv), {}).get(sel) if ah2_lv is not None else None

    def g_2h_ou(sel):
        return hou.get((mk, "_2H", ou2_lv), {}).get(sel) if ou2_lv is not None else None

    # --- 盘口块 ---
    base_r = r  # 记录盘口块起始行(主队行), 用于之后标红正确选项
    # 主队行
    C(1, home)
    C(3, fmt_1x2("home", g_ft_x2("home")))
    C(4, fmt_ah("home", ah_lv, g_ft_ah("home")))
    C(5, fmt_ou("over", ou_lv, g_ft_ou("over")))
    C(7, fmt_1x2("home", g_2h_x2("home")))
    C(8, fmt_ah("home", ah2_lv, g_2h_ah("home")))
    C(9, fmt_ou("over", ou2_lv, g_2h_ou("over")))
    C(10, ht_home)          # K列: 半场比分
    C(11, ft_home)          # L列: 终场比分
    r += 1
    # 客队行
    C(1, away)
    C(3, fmt_1x2("away", g_ft_x2("away")))
    C(4, fmt_ah("away", ah_lv, g_ft_ah("away")))
    C(5, fmt_ou("under", ou_lv, g_ft_ou("under")))
    C(7, fmt_1x2("away", g_2h_x2("away")))
    C(8, fmt_ah("away", ah2_lv, g_2h_ah("away")))
    C(9, fmt_ou("under", ou2_lv, g_2h_ou("under")))
    C(10, ht_away)          # K列: 半场比分
    C(11, ft_away)          # L列: 终场比分
    r += 1
    # 和局行
    C(3, fmt_1x2("draw", g_ft_x2("draw")))
    C(7, fmt_1x2("draw", g_2h_x2("draw")))
    r += 1

    # --- 全场盘口正确选项标红 (让球 / 大小球) ---
    # 仅在有终场比分时判定; 走水(正好等于盘口线)不标红.
    # 仅当该方向存在真实赔率(>0, 即显示非 '--')时才标红, 否则只显示 '--' 无可标注选项.
    if has_outcome and ft_hit:
        sh_i, sa_i = int(sh), int(sa)
        # 让球: 盘口线 lv 含义为 "主队 - lv vs 客队"; diff = 主进球 - 客进球 - lv
        if ah_lv is not None:
            diff = sh_i - sa_i - ah_lv
            if diff > 0 and (g_ft_ah("home") or 0) > 0:
                ws.cell(base_r, 5).font = red        # 主队赢盘 -> 主队行(让球列 E)
            elif diff < 0 and (g_ft_ah("away") or 0) > 0:
                ws.cell(base_r + 1, 5).font = red    # 客队赢盘 -> 客队行(让球列 E)
        # 大小球: total vs 线; 大 -> 主队行(写"大"的列 F), 小 -> 客队行(写"小"的列 F)
        if ou_lv is not None:
            total = sh_i + sa_i
            if total > ou_lv and (g_ft_ou("over") or 0) > 0:
                ws.cell(base_r, 6).font = red        # 大球 -> 主队行(大小列 F)
            elif total < ou_lv and (g_ft_ou("under") or 0) > 0:
                ws.cell(base_r + 1, 6).font = red    # 小球 -> 客队行(大小列 F)

    # --- 全场1X2正确结果标红 (主胜/和局/客胜) ---
    # 这是"比赛正确结果"最核心的标注: 有终场比分即标, 不依赖赔率是否存在.
    if has_outcome and ft_hit:
        sh_i, sa_i = int(sh), int(sa)
        if sh_i > sa_i:
            ws.cell(base_r, 4).font = red          # 主胜 -> 主队行(全场独赢列 D)
        elif sh_i < sa_i:
            ws.cell(base_r + 1, 4).font = red      # 客胜 -> 客队行(全场独赢列 D)
        else:
            ws.cell(base_r + 2, 4).font = red      # 和局 -> 和局行(全场独赢列 D)
        # 终场比分(主客各自行 L 列)标红 -> 直接高亮"结果比分"
        ws.cell(base_r, 12).font = red
        ws.cell(base_r + 1, 12).font = red

    # --- 半场比分标红 (仅确认型半场: 实测/回填; 不依赖全场赛果) ---
    if has_ht and ht_confirmed:
        ws.cell(base_r, 11).font = red         # 主队半场进球 -> 主队行(半场比分列 K)
        ws.cell(base_r + 1, 11).font = red     # 客队半场进球 -> 客队行(半场比分列 K)

    # --- 波胆块 (每场比赛固定渲染, 无波胆则填 '--') ---
    C(1, "初盘波胆比分")
    C(4, "上半场波胆")
    C(7, "下半场波胆")
    r += 1
    C(1, home); C(2, "全场平局"); C(3, away)
    C(4, home); C(5, "上半场平局"); C(6, away)
    C(7, home); C(8, "下半场平局"); C(9, away)
    r += 1

    def _hcs_cell(suf, side, template_score):
        """上半场/下半场波胆赔率; 数据源有则显真实赔率, 无则 '--'."""
        if not template_score:
            return ""
        score = "其他" if template_score == "其他" else template_score.replace(":", "-")
        odds = hcs.get((mk, suf, side, score))
        if not odds or odds <= 0:
            return f"{template_score}      --"
        return f"{template_score}      {odds:.2f}"

    for hs, ds, as_ in CS_ROWS:
        # 初盘全场波胆: GQ 的 CS 市场用户确认为下半场/剩余时间波胆, 此处全场初盘留 '--'
        C(1, f"{hs}      --" if hs else "")
        C(2, f"{ds}      --" if ds else "")
        C(3, f"{as_}      --" if as_ else "")
        # 上半场波胆: GQ 无 CS_1H 市场
        C(4, _hcs_cell("_1H", "home", hs))
        C(5, _hcs_cell("_1H", "draw", ds))
        C(6, _hcs_cell("_1H", "away", as_))
        # 下半场波胆: 使用 GQ CS 市场数据
        C(7, _cs_cell(cs, mk, "home", hs))
        C(8, _cs_cell(cs, mk, "draw", ds))
        C(9, _cs_cell(cs, mk, "away", as_))

        # 命中红字: 下半场波胆格 == 下半场比分(有半场比分时可算)
        if ft_hit and ht_hit and ht_h is not None and ht_a is not None:
            sh2 = f"{int(sh) - int(ht_h)}:{int(sa) - int(ht_a)}"
            if hs == sh2: ws.cell(r, 8).font = red
            if ds == sh2: ws.cell(r, 9).font = red
            if as_ == sh2: ws.cell(r, 10).font = red

        r += 1
    r += 1  # 每场之间空行
    return r


# ---------------- 主流程 ----------------
def main():
    ap = argparse.ArgumentParser(description="GQ 数据 -> 用户模板 Excel")
    ap.add_argument("--league", default=None, help="仅导出指定联赛(模糊匹配)")
    ap.add_argument("--limit", type=int, default=0, help="限制导出场数(0=全量)")
    ap.add_argument("--db", default=DB)
    ap.add_argument("--out", default=OUT)
    args = ap.parse_args()

    import openpyxl
    c = sqlite3.connect(args.db)
    cur = c.cursor()

    # 终场比分: 优先 match_outcomes(官方归档), 缺失时回退 matches.score_home
    #   —— 关键修复: 采集器只把少量比赛归档进 match_outcomes, 但 matches 表里
    #      有 1803 场 finished 带着真实比分却没进 outcomes, 导致正确结果被渲染成 0.0.
    # 联赛: 优先 matches.league, 空时回退 match_outcomes.league.
    # 排序: 空联赛排到末尾, 避免文件顶部一堆"联赛未标注"误导.
    sql = """
    SELECT m.match_key, m.home, m.away,
           COALESCE(NULLIF(m.league, ''), o.league) AS league, m.status,
           m.score_home AS m_h, m.score_away AS m_a,
           COALESCE(o.score_home, CASE WHEN m.status='finished' THEN m.score_home END) AS sh,
           COALESCE(o.score_away, CASE WHEN m.status='finished' THEN m.score_away END) AS sa,
           COALESCE(o.ht_score_home, m.ht_score_home) AS ht_h,
           COALESCE(o.ht_score_away, m.ht_score_away) AS ht_a
    FROM matches m
    LEFT JOIN match_outcomes o ON o.mid = m.mid
    WHERE 1=1
    """
    params = []
    if args.league:
        sql += " AND COALESCE(NULLIF(m.league, ''), o.league) LIKE ?"
        params.append(f"%{args.league}%")
    sql += """ ORDER BY
        (CASE WHEN COALESCE(NULLIF(m.league, ''), o.league) IS NULL THEN 1 ELSE 0 END),
        COALESCE(NULLIF(m.league, ''), o.league),
        m.kickoff DESC"""
    matches = cur.execute(sql, params).fetchall()

    ft, cs, ft_ah_line, ft_ou_line = load_opening(cur)
    h1x2, hah, hou, hcs, hah_line, hou_line = load_half(cur)
    c.close()

    if args.limit:
        matches = matches[: args.limit]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # 列宽: 拉到能完整显示(避免赔率被列边界切掉/遮挡字体)
    #   盘口 AH/OU 用 6 空格对齐 -> 需更宽; 联赛名可能很长 -> 最宽
    _colw = {'A': 42, 'B': 28, 'C': 14, 'D': 28, 'E': 28,
             'F': 18, 'G': 14, 'H': 28, 'I': 18, 'J': 28,
             'K': 14, 'L': 14}
    for _col, _w in _colw.items():
        ws.column_dimensions[_col].width = _w

    r = 1
    cur_league = None
    exported = 0
    cs_mks = set(k[0] for k in cs)
    for mk, home, away, league, status, m_h, m_a, sh, sa, ht_h, ht_a in matches:
        # 不再跳过任何比赛: 全部渲染(含未开赛/无赔率场次, 其结果为空/0.0不标红).
        # 旧逻辑: 仅导出 有赛果/有盘口/live; 现改为全量, 避免"遗漏比赛".
        league_disp = league or "联赛未标注"
        if league_disp != cur_league:
            ws.cell(r, 1, league_disp)
            ws.cell(r, 2, "初盘赔率")
            ws.cell(r, 4, "全场独赢")
            ws.cell(r, 5, "全场让球")
            ws.cell(r, 6, "全场大小")
            ws.cell(r, 8, "下半场赔率")
            ws.cell(r, 11, "半场比分")
            ws.cell(r, 12, "终场比分")
            r += 1
            cur_league = league_disp
        has_outcome = sh is not None and sa is not None
        r = render_match_block(ws, r, mk, home, away, status, m_h, m_a, ht_h, ht_a,
                               has_outcome, sh, sa,
                               ft, cs, ft_ah_line, ft_ou_line,
                               h1x2, hah, hou, hcs, hah_line, hou_line)
        exported += 1

    wb.save(args.out)
    n_2h_1x2 = sum(1 for (mk, suf) in h1x2 if suf == "_2H")
    n_1h_cs = sum(1 for (mk, suf, side, score) in hcs if suf == "_1H")
    print(f"导出完成: {exported} 场, {r - 1} 行 -> {args.out}")
    print(f"  全场盘口记录: {len(ft)}  波胆记录: {len(cs)}  半场波胆记录: {n_1h_cs}")
    print(f"  下半场 1X2 比赛数: {n_2h_1x2}  AH_2H 线记录: {len(hah)}  OU_2H 线记录: {len(hou)}")


if __name__ == "__main__":
    main()
